"""Import mice from existing lab colony spreadsheets (.xlsx).

Real lab sheets vary: columns appear in different orders, some tabs aren't
colonies at all, and cells contain stray values like "/", "#REF!", or blanks.
So this importer is **header-driven** — it finds the header row and maps
columns by their *name*, not their position — and it tolerates messy data.

Typical recognized headers (case/spacing/“?”/“#” insensitive):

    cage, number, sex, dob, dod, age, geno, tags?, breeder pair[ #],
    PARENT PAIR, generation, use, link, notes

The public entry points are:
    parse_workbook(path)            -> a Preview describing each sheet
    import_preview(preview, names)  -> writes the chosen sheets to the database
"""
from datetime import date, datetime

from openpyxl import load_workbook

from . import db
from .models import Cage, Colony, Mouse

# Map a normalized header -> the Mouse field it fills. Two sentinels are
# handled specially: __cage__ (the cage label) and __generation__ (folded
# into notes). Headers not listed here (age, today's date, blanks) are ignored.
HEADER_MAP = {
    "cage": "__cage__",
    "number": "tag", "num": "tag", "mouse": "tag", "id": "tag", "no": "tag",
    "sex": "sex",
    "dob": "dob", "d.o.b": "dob", "birth": "dob", "date of birth": "dob",
    "dod": "date_of_death", "d.o.d": "date_of_death", "death": "date_of_death",
    "date of death": "date_of_death",
    "geno": "genotype", "genotype": "genotype", "gene": "genotype",
    "tags": "ear_tags", "tag": "ear_tags", "ear tags": "ear_tags",
    "ear tag": "ear_tags",
    "breeder pair": "breeder_pair", "breeder": "breeder_pair",
    "parent pair": "parent_pair", "parents": "parent_pair",
    "generation": "__generation__",
    "use": "use",
    "link": "link",
    "notes": "notes", "note": "notes",
}

# A sheet is treated as a colony only if its header row provides at least these.
REQUIRED_TARGETS = {"sex", "genotype"}
ID_TARGETS = {"__cage__", "tag"}  # plus at least one of these

MAX_BLANK_RUN = 50  # stop scanning a sheet after this many empty rows in a row


def _norm(value):
    """Normalize a header cell: lowercase, trimmed, without trailing ? or #."""
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = text.strip("?#").strip()
    return " ".join(text.split())


def _as_label(value):
    """Render a cage/number cell as a clean string (no trailing '.0')."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, (int,)):
        return str(value)
    return str(value).strip()


def _as_date(value):
    """Return a date for real date cells, else None (handles '/', '#REF!')."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _as_sex(value):
    text = str(value).strip().lower() if value is not None else ""
    if text.startswith("f"):
        return "F"
    if text.startswith("m"):
        return "M"
    return "U"


def _as_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


class SheetResult:
    """What we found on one worksheet."""

    def __init__(self, sheet_name):
        self.sheet_name = sheet_name
        self.colony_name = sheet_name.strip()
        self.is_colony = False
        self.skip_reason = ""
        self.mice = []          # list of dicts of Mouse fields
        self.cage_labels = set()

    @property
    def mouse_count(self):
        return len(self.mice)


class Preview:
    """The result of parsing a whole workbook."""

    def __init__(self, path):
        self.path = path
        self.sheets = []  # list[SheetResult]

    @property
    def colony_sheets(self):
        return [s for s in self.sheets if s.is_colony]

    @property
    def skipped_sheets(self):
        return [s for s in self.sheets if not s.is_colony]


def _find_header_row(ws):
    """Return (row_index, {column_index: target}) for the best header row."""
    best = None
    for r in range(1, min(ws.max_row, 15) + 1):
        mapping = {}
        for c in range(1, ws.max_column + 1):
            target = HEADER_MAP.get(_norm(ws.cell(row=r, column=c).value))
            if target:
                mapping[c] = target
        targets = set(mapping.values())
        if REQUIRED_TARGETS.issubset(targets) and (targets & ID_TARGETS):
            score = len(mapping)
            if best is None or score > best[2]:
                best = (r, mapping, score)
    if best is None:
        return None, None
    return best[0], best[1]


def _parse_sheet(ws):
    result = SheetResult(ws.title)
    header_row, mapping = _find_header_row(ws)
    if header_row is None:
        result.skip_reason = "no recognizable colony headers"
        return result

    result.is_colony = True
    blank_run = 0
    for r in range(header_row + 1, ws.max_row + 1):
        row = {target: ws.cell(row=r, column=c).value
               for c, target in mapping.items()}
        cage = _as_label(row.get("__cage__"))
        tag = _as_label(row.get("tag"))
        if not cage and not tag:
            blank_run += 1
            if blank_run >= MAX_BLANK_RUN:
                break
            continue
        blank_run = 0

        notes = _as_text(row.get("notes"))
        generation = _as_text(row.get("__generation__"))
        if generation:
            notes = (notes + " | " if notes else "") + f"Generation: {generation}"

        dod = _as_date(row.get("date_of_death"))
        mouse = {
            "cage_label": cage,
            "tag": tag or "(unknown)",
            "sex": _as_sex(row.get("sex")),
            "dob": _as_date(row.get("dob")),
            "date_of_death": dod,
            "status": "dead" if dod else "alive",
            "genotype": _as_text(row.get("genotype")),
            "ear_tags": _as_text(row.get("ear_tags")),
            "breeder_pair": _as_text(row.get("breeder_pair")),
            "parent_pair": _as_text(row.get("parent_pair")),
            "use": _as_text(row.get("use")),
            "link": _as_text(row.get("link")),
            "notes": notes,
        }
        result.mice.append(mouse)
        if cage:
            result.cage_labels.add(cage)
    if not result.mice:
        result.is_colony = False
        result.skip_reason = "looks like a colony sheet but has no data rows"
    return result


def parse_workbook(path):
    """Read a workbook and describe what would be imported (no DB writes)."""
    preview = Preview(path)
    wb = load_workbook(path, data_only=True, read_only=True)
    for ws in wb.worksheets:
        preview.sheets.append(_parse_sheet(ws))
    wb.close()
    return preview


def _get_or_make_cage(colony_id, label, cages):
    """Return the Cage for a label within a colony, creating it once."""
    if not label:
        return None
    if label not in cages:
        cage = Cage(colony_id=colony_id, label=label)
        db.session.add(cage)
        db.session.flush()
        cages[label] = cage
    return cages[label]


def _add_mouse(colony_id, cages, m, source_sheet=None):
    notes = m["notes"]
    if source_sheet:
        notes = (notes + " | " if notes else "") + f"Source: {source_sheet}"
    cage = _get_or_make_cage(colony_id, m["cage_label"], cages)
    db.session.add(Mouse(
        colony_id=colony_id,
        cage_id=cage.id if cage else None,
        tag=m["tag"],
        sex=m["sex"],
        dob=m["dob"],
        date_of_death=m["date_of_death"],
        status=m["status"],
        genotype=m["genotype"],
        ear_tags=m["ear_tags"],
        breeder_pair=m["breeder_pair"],
        parent_pair=m["parent_pair"],
        use=m["use"],
        link=m["link"],
        notes=notes,
    ))


def import_preview(preview, selected_sheet_names, combine=False,
                   combined_name=None):
    """Write the chosen sheets to the database. Returns a list of summaries.

    If ``combine`` is True, every selected sheet is loaded into a *single*
    colony named ``combined_name`` (the originating tab is recorded in each
    mouse's notes). Otherwise each sheet becomes its own colony.
    """
    sheets = [s for s in preview.colony_sheets
              if s.sheet_name in selected_sheet_names]
    if not sheets:
        return []

    if combine:
        colony = Colony(
            name=combined_name or "Imported colony",
            description="Imported from spreadsheet "
                        f"({', '.join(s.sheet_name for s in sheets)}).",
        )
        db.session.add(colony)
        db.session.flush()
        cages = {}
        multi = len(sheets) > 1
        for sheet in sheets:
            for m in sheet.mice:
                _add_mouse(colony.id, cages,
                           m, sheet.sheet_name if multi else None)
        db.session.commit()
        return [{
            "colony_id": colony.id,
            "colony_name": colony.name,
            "mice": sum(s.mouse_count for s in sheets),
            "cages": len(cages),
        }]

    summaries = []
    for sheet in sheets:
        colony = Colony(
            name=sheet.colony_name or "Imported colony",
            description=f"Imported from spreadsheet sheet “{sheet.sheet_name}”.",
        )
        db.session.add(colony)
        db.session.flush()
        cages = {}
        for m in sheet.mice:
            _add_mouse(colony.id, cages, m)
        summaries.append({
            "colony_id": colony.id,
            "colony_name": colony.name,
            "mice": sheet.mouse_count,
            "cages": len(cages),
        })
    db.session.commit()
    return summaries
