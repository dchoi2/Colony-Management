"""Generate a nicely formatted Excel (.xlsx) workbook of the colony.

The layout mirrors a typical lab colony spreadsheet (cage, number, sex,
dates, genotype, tags, breeding info, use, link) but with consistent
formatting: a styled header, frozen top row, filter dropdowns, sensible
column widths, real dates, and deceased animals flagged in red.
"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --- styling -------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="2E7D32")   # green, like the sample
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL = PatternFill("solid", fgColor="F2F7F2")      # subtle zebra striping
DECEASED_FONT = Font(color="B00020")                   # dark red for dead/removed
_THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

# Each column: (header, value getter, width, alignment kind)
COLUMNS = [
    ("Cage", lambda m: m.cage.label if m.cage else "", 9, "center"),
    ("Num", lambda m: m.tag, 8, "center"),
    ("Sex", lambda m: m.sex.lower() if m.sex in ("M", "F") else "", 6, "center"),
    ("DOB", lambda m: m.dob, 12, "date"),
    ("DOD", lambda m: m.date_of_death, 12, "date"),
    ("Age (wk)", lambda m: m.age_weeks, 9, "center"),
    ("Geno", lambda m: m.genotype, 10, "center"),
    ("Strain", lambda m: m.strain, 14, "left"),
    ("Tags", lambda m: m.ear_tags, 9, "center"),
    ("Breeder Pair #", lambda m: m.breeder_pair, 14, "center"),
    ("Parent Pair", lambda m: m.parent_pair, 16, "left"),
    ("Use", lambda m: m.use, 22, "left"),
    ("Link", lambda m: m.link, 28, "left"),
    ("Status", lambda m: m.status, 10, "center"),
    ("Colony", lambda m: m.colony.name if m.colony else "", 18, "left"),
    ("Notes", lambda m: m.notes, 30, "left"),
]


def _numeric_key(tag):
    """Sort tags numerically when they look like numbers, else alphabetically."""
    try:
        return (0, int(tag))
    except (TypeError, ValueError):
        return (1, str(tag))


def sort_mice(mice):
    """Group by cage, then by animal number — like a hand-kept colony sheet."""
    return sorted(
        mice,
        key=lambda m: (m.cage.label if m.cage else "~", _numeric_key(m.tag)),
    )


def build_mice_workbook(mice, sheet_title="Colony"):
    """Return a BytesIO containing a formatted .xlsx of the given mice."""
    mice = sort_mice(mice)

    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_title or "Colony")[:31]

    # Header row
    for col_idx, (header, _getter, width, _align) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 20

    # Data rows
    for row_idx, m in enumerate(mice, start=2):
        deceased = bool(m.date_of_death) or (m.status and m.status != "alive")
        striped = row_idx % 2 == 0
        for col_idx, (_header, getter, _width, align) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=getter(m))
            cell.border = BORDER
            if align == "date":
                cell.number_format = "mm/dd/yyyy"
                cell.alignment = CENTER
            else:
                cell.alignment = CENTER if align == "center" else LEFT
            if striped:
                cell.fill = ALT_FILL
            if deceased:
                cell.font = DECEASED_FONT

    # Freeze the header and add filter dropdowns over every column.
    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col}{max(1, len(mice) + 1)}"

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
